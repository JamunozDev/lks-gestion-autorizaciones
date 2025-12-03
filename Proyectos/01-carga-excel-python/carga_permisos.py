import signal, sys
from openpyxl import load_workbook
import sqlite3


def cortarPrograma(sig, frame):
    print("\n\n[!]Saliendo ...\n")
    sys.exit(1)


# Ctlr + C
signal.signal(signal.SIGINT, cortarPrograma)

# Variables globales
fichero_permisos = "docs/excel-cod.xlsx"
lista_permisos = []


conn = sqlite3.connect("db/permisos.db")  # crea la BD si no existe
cursor = conn.cursor()


def cargaPermisos():
    print("Iniciando carga de permisos")

    wb = load_workbook(fichero_permisos)
    for nombre_hoja in wb.sheetnames:
        hoja = wb[nombre_hoja]
        for fila in hoja.iter_rows(min_row=2, values_only=True):
            codigo = fila[10]
            descripcion = fila[3]
            if codigo and not any(u["codigo"] == codigo for u in lista_permisos):
                lista_permisos.append(
                    {
                        "codigo": codigo,
                        "descripcion": descripcion,
                        "formulario": nombre_hoja,
                    }
                )

    cursor.execute("delete from permisos")
    conn.commit()
    cursor.executemany(
        "INSERT INTO permisos (codigo, descripcion, formulario) VALUES (:codigo, :descripcion, :formulario)",
        lista_permisos,
    )
    conn.commit()


if __name__ == "__main__":

    cargaPermisos()
